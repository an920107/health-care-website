# -*- coding: utf-8 -*-

import os
import uuid
from io import BytesIO
from pathlib import Path
from config import Config
from datetime import datetime

from script.utils import api_input_check, api_input_get
from models.models import RestaurantPost, db, RestaurantAttachment
from models.responses import Response

import pandas as pd
from openpyxl import Workbook
from flask import Blueprint, request, send_file

restaurant_post_blueprint = Blueprint('restaurant_post', __name__)


@restaurant_post_blueprint.route('/<int:post_id>', methods=['GET'])
def get_restaurant_post(post_id):
    """
    Get restaurant_post post by id
    ---
    tags:
      - Restaurant Post
    parameters:
      - name: post_id
        in: path
        type: integer
        required: true
        description: post id
    responses:
      200:
        description: get post success
      404:
        description: post not found
    """
    restaurant_post = db.session.query(RestaurantPost).get(post_id)
    if restaurant_post is None:
        return Response.not_found('post not found')
    return Response.response('get post success', restaurant_post.as_dict())


@restaurant_post_blueprint.route('', methods=['GET'])
def get_restaurant_posts():
    """
    Get restaurant_post posts
    ---
    tags:
      - Restaurant Post
    parameters:
      - name: page
        in: query
        type: integer
        required: false
        description: page
    responses:
      200:
        description: get post success
      404:
        description: post not found
    """
    posts = db.session.query(RestaurantPost).all()

    if request.args.get('page') and int(request.args['page']) >= 1:
        page = int(request.args['page'])
    else:
        page = 1

    posts = posts[(page - 1) * Config.PAGE_SIZE:page * Config.PAGE_SIZE]

    payload = {
        'total_page': str(len(posts) // Config.PAGE_SIZE + 1),
        'posts': [post.as_dict() for post in posts],
        'page': str(page)
    }
    return Response.response('get post success', payload)


@restaurant_post_blueprint.route('', methods=['POST'])
def post_restaurant_post():
    """
    Post restaurant_post post
    ---
    tags:
      - Restaurant Post
    parameters:
      - name: title
        in: formData
        type: string
        required: true
        description: title
      - name: attachments
        in: formData
        type: string
        required: true
        description: attachments
      - name: category
        in: formData
        type: string
        required: true
        description: category
      - name: time
        in: formData
        type: string
        required: true
        description: time
      - name: valid
        in: formData
        type: string
        required: true
        description: valid
    responses:
      200:
        description: post restaurant_post success
    """
    if not api_input_check(['title', 'attachments', 'category', 'time', 'valid'], request.form):
        return Response.client_error("missing ['title', 'attachments', 'category', 'time', 'valid'] form data")

    title, attachments, category, time, valid = api_input_get(
        ['title', 'attachments', 'category', 'time', 'valid'], request.form
    )

    restaurant_post = RestaurantPost(
        title=title, attachments=attachments, category=category, time=datetime.strptime(time, "%Y-%m-%dT%H:%M:%S.%f"),
        valid=valid
    )
    db.session.add(restaurant_post)
    db.session.commit()
    return Response.response('post restaurant_post success', restaurant_post.as_dict())


@restaurant_post_blueprint.route('/<int:post_id>', methods=['PUT'])
def put_restaurant_post(post_id):
    """
    Put restaurant_post post by id
    ---
    tags:
      - Restaurant Post
    parameters:
      - name: post_id
        in: path
        type: integer
        required: true
        description: post id
      - name: title
        in: formData
        type: string
        required: true
        description: title
      - name: attachments
        in: formData
        type: string
        required: true
        description: attachments
      - name: category
        in: formData
        type: string
        required: true
        description: category
      - name: time
        in: formData
        type: string
        required: true
        description: time
      - name: valid
        in: formData
        type: string
        required: true
        description: valid
    responses:
      200:
        description: put restaurant_post success
      404:
        description: post not found
    """
    restaurant_post = db.session.query(RestaurantPost).get(post_id)
    if restaurant_post is None:
        return Response.not_found('post not found')

    if not api_input_check(['title', 'attachments', 'category', 'time', 'valid', 'visible'], request.form):
        return Response.client_error(
            "missing ['title', 'attachments', 'category', 'time', 'valid', 'visible'] form data")

    title, attachments, category, time, valid, visible = api_input_get(
        ['title', 'attachments', 'category', 'time', 'valid', 'visible'], request.form
    )

    restaurant_post.title = title
    restaurant_post.attachments = attachments
    restaurant_post.category = category
    restaurant_post.time = datetime.strptime(time, "%Y-%m-%dT%H:%M:%S.%f")
    restaurant_post.valid = valid
    restaurant_post.visible = visible
    db.session.commit()

    return Response.response('put restaurant_post success', restaurant_post.as_dict())


@restaurant_post_blueprint.route('/<int:post_id>', methods=['DELETE'])
def delete_restaurant_post(post_id):
    """
    Delete restaurant_post post by id
    ---
    tags:
      - Restaurant Post
    parameters:
      - name: post_id
        in: path
        type: integer
        required: true
        description: post id
    responses:
      200:
        description: delete restaurant_post success
      404:
        description: post not found
    """
    restaurant_post = RestaurantPost.query.get(post_id)
    if restaurant_post is None:
        return Response.not_found('post not found')

    attachments = RestaurantAttachment.query.filter_by(post_id=post_id).all()
    for attachment in attachments:
        os.remove(attachment.file_path)
        db.session.delete(attachment)

    db.session.delete(restaurant_post)
    db.session.commit()
    return Response.response('delete restaurant_post success')


@restaurant_post_blueprint.route('/<int:post_id>/visible', methods=['PATCH'])
def patch_restaurant_post_visible(post_id):
    """
    Patch restaurant_post post visible
    ---
    tags:
      - Restaurant Post
    parameters:
      - name: post_id
        in: path
        type: integer
        required: true
        description: post id
      - name: visible
        in: formData
        type: string
        required: true
        description: visible
    responses:
      200:
        description: patch restaurant_post visible success
      404:
        description: post not found
    """
    restaurant_post = db.session.query(RestaurantPost).get(post_id)
    if restaurant_post is None:
        return Response.not_found('post not found')

    visible = request.form.get('visible')
    restaurant_post.visible = visible
    db.session.commit()
    return Response.response('patch restaurant_post visible success', restaurant_post.as_dict())


@restaurant_post_blueprint.route('/<int:post_id>/attachment', methods=['POST'])
def post_restaurant_attachment(post_id):
    """
    Post restaurant_post attachment
    ---
    tags:
      - Restaurant Post
    parameters:
      - name: post_id
        in: path
        type: integer
        required: true
        description: post id
      - name: blob_attachment
        in: formData
        type: file
        required: true
        description: The attachment to upload
    responses:
      200:
        description: post attachment successful
      400:
        description: no blob attachment
    """
    if 'blob_attachment' not in request.files:
        return Response.client_error('no blob attachment')

    blob_attachment = request.files['blob_attachment']
    file_name = f"{uuid.uuid4()}.{blob_attachment.filename.split('.')[-1]}."
    file_path = Path(Config.RESTAURANT_CONFIG['ATTACHMENT_DIR']) / Path(file_name)
    blob_attachment.save(file_path)

    attachment = RestaurantAttachment(name=blob_attachment.filename, file_path=str(file_path), post_id=post_id)
    db.session.add(attachment)
    db.session.commit()
    return Response.response(
        'add attachments successful',
        {
            'attachment_id': str(attachment.id),
            'attachment_name': attachment.name,
            'attachment_uri': f'/api/attachment/restaurant_post/{attachment.id}',
            'attachment_info': f'/api/attachment/restaurant_post/{attachment.id}/info'
        })


@restaurant_post_blueprint.route('stats', methods=['GET'])
def get_restaurant_stats():
    """
    Get restaurant_post stats
    ---
    tags:
      - Restaurant Post
    parameters:
      - name: start_date
        in: query
        type: string
        required: true
        description: start date
      - name: end_date
        in: query
        type: string
        required: true
        description: end date
    responses:
      200:
        description: get restaurant_post stats success
      400:
        description: missing ['start_date', 'end_date'] query data
    """
    if not api_input_check(['start_date', 'end_date'], request.args):
        return Response.client_error("missing ['start_date', 'end_date'] query data")

    start_date, end_date = api_input_get(['start_date', 'end_date'], request.args)
    start_date = datetime.strptime(start_date, "%Y-%m-%d")
    end_date = datetime.strptime(end_date, "%Y-%m-%d")
    posts = RestaurantPost.query.filter(
        RestaurantPost.time >= start_date, RestaurantPost.time <= end_date
    ).all()

    category_validation_count = {
        'water': {'total': 0, '1': 0, '0': 0, 'valid_rate': 0},
        'food': {'total': 0, '1': 0, '0': 0, 'valid_rate': 0},
        'drink': {'total': 0, '1': 0, '0': 0, 'valid_rate': 0},
        'ice': {'total': 0, '1': 0, '0': 0, 'valid_rate': 0},
        'others': {'total': 0, '1': 0, '0': 0, 'valid_rate': 0},
        'total': {'total': 0, '1': 0, '0': 0, 'valid_rate': 0}
    }

    for post in posts:
        category_validation_count[post.category][post.valid] += 1
        category_validation_count[post.category]['total'] += 1
        category_validation_count['total'][post.valid] += 1
        category_validation_count['total']['total'] += 1

    for category, count in category_validation_count.items():
        if count['total'] != 0:
            count['valid_rate'] = count['1'] / count['total']

    df = pd.DataFrame(category_validation_count).T
    df.columns = ['總件數', '合格幾件', '不合格幾件', '合格率']
    df['類別'] = ['飲用水', '熟食', '飲料', '冰塊', '其他', '總和']
    df = df[['類別', '總件數', '合格幾件', '不合格幾件', '合格率']]

    wb = Workbook()
    ws = wb.active
    column_map = {0: 'A', 1: 'B', 2: 'C', 3: 'D', 4: 'E'}
    for col_idx, col_name in column_map.items():
        ws[col_name + '1'] = df.columns[col_idx]

    for row_idx in range(len(df)):
        for col_idx, col_name in column_map.items():
            ws[col_name + str(row_idx + 2)] = df.iloc[row_idx, col_idx]

    ws['A8'] = '開始日期'
    ws['B8'] = start_date.strftime('%Y-%m-%d')
    ws['C8'] = '結束日期'
    ws['D8'] = end_date.strftime('%Y-%m-%d')
    ws['A9'] = '餐廳檢查報告統計表'

    excel_data = BytesIO()
    wb.save(excel_data)
    excel_data.seek(0)

    return send_file(
        excel_data,
        as_attachment=True,
        download_name='餐廳檢查報告.xlsx',
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )

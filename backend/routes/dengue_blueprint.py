# -*- coding: utf-8 -*-
import json
from models.models import Building, db, Dengue

import pandas as pd
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from io import BytesIO

from script.utils import api_input_check, api_input_get
from script.dengue_scripts import DengueDataClass
from datetime import datetime, timedelta
from script.oauth_scripts import authorization_required

from models.models import StaticPost, db, StaticImage, StaticAttachment
from models.responses import Response

from flask import Blueprint, request, send_file
from flask_jwt_extended import get_jwt_identity

dengue_blueprint = Blueprint('dengue', __name__)

to_month = lambda x: int(x.split('-')[0]) * 12 + int(x.split('-')[1])
to_year_month = lambda x: f"{x // 12}-{x % 12 + 1:02d}"


@dengue_blueprint.route('/building', methods=['GET'])
def get_buildings():
    """
    Get all buildings
    ---
    tags:
      - building
    parameters:
      - name: user_id
        in: query
        type: string
        required: false
        description: The id of the user
    responses:
      200:
        description: Get all buildings
    """
    rules = {}
    if api_input_check(['user_id'], request.args):
        rules['user_id'] = api_input_get(['user_id'], request.args)[0]
    buildings = Building.query.filter_by(**rules).all()
    return Response.response('get buildings successful', [building.as_dict() for building in buildings])


@dengue_blueprint.route('/building/<int:building_id>', methods=['GET'])
def get_building(building_id):
    """
    Get single buildings
    ---
    tags:
      - building
    parameters:
      - name: building_id
        in: path
        type: string
        required: true
        description: The id of the user
    responses:
      200:
        description: Get all buildings
    """
    building = Building.query.get(building_id)
    return Response.response('get buildings successful', building.as_dict())


@dengue_blueprint.route('/building', methods=['POST'])
@authorization_required([0, 1, 2])
def create_building():
    """
    create_building
    ---
    tags:
      - building
    parameters:
      - name: chinese_name
        in: formData
        type: string
        required: true
        description: The chinese name of the building
      - name: user_id
        in: formData
        type: string
        required: true
        description: The id of the user
    """
    if not api_input_check(['chinese_name', 'user_id'], request.form):
        return Response.client_error("no ['chinese_name', 'user_id'] in form")

    chinese_name, user_id = api_input_get(['chinese_name', 'user_id'], request.form)
    building = Building(chinese_name=chinese_name, user_id=user_id)
    db.session.add(building)
    db.session.commit()
    return Response.response('create building successful', building.as_dict())


@dengue_blueprint.route('/building/<int:building_id>', methods=['PATCH'])
@authorization_required([0, 1, 2])
def update_building(building_id):
    """
    update a building
    ---
    tags:
      - building
    parameters:
      - name: chinese_name
        in: formData
        type: string
        required: true
        description: The chinese name of the building
    """
    if not api_input_check(['user_id'], request.form):
        return Response.client_error("no ['user_id'] in form")

    user_id, = api_input_get(['user_id'], request.form)
    building = Building.query.get(building_id)
    building.user_id = user_id
    db.session.commit()
    return Response.response('update building successful', building.as_dict())


@dengue_blueprint.route('/building/<int:building_id>', methods=['DELETE'])
@authorization_required([0, 1, 2])
def delete_building(building_id):
    """
    Create a building
    ---
    tags:
      - building
    parameters:
      - name: building_id
        in: path
        type: integer
        required: true
        description: The id of the building
    """
    building = Building.query.filter_by(id=building_id).first()
    if building is None:
        return Response.client_error("delete building successful")
    db.session.delete(building)
    db.session.commit()
    return Response.response('delete building successful')


@dengue_blueprint.route('/form', methods=['POST'])
@authorization_required()
def create_form():
    """
    Create a form
    ---
    tags:
      - Dengue
    parameters:
      - name: json_data
        in: formData
        type: string
        required: true
        description: The json data of the form
      - name: building
        in: formData
        type: string
        required: true
        description: The json data of the form
      - name: create_year_month
        in: formData
        type: string
        required: true
        description: start date
    """

    if not api_input_check(['json_data', 'create_year_month', 'building'], request.form):
        return Response.client_error("no ['json_data', 'create_year_month', 'building'] in form")

    json_data, create_year_month, building_id = api_input_get(['json_data', 'create_year_month', 'building'],
                                                              request.form)
    dengue = Dengue(json_data=json_data, create_year_month=create_year_month, building_id=building_id)
    db.session.add(dengue)
    db.session.commit()
    return Response.response('create form successful', dengue.as_dict())


@dengue_blueprint.route('/form/<int:form_id>', methods=['DELETE'])
@authorization_required()
def delete_form(form_id):
    """
    Delete a form
    ---
    tags:
      - Dengue
    parameters:
      - name: form_id
        in: path
        type: integer
        required: true
        description: The id of the form
    """
    dengue = Dengue.query.get(form_id)
    db.session.delete(dengue)
    db.session.commit()
    return Response.response('delete form successful')


@dengue_blueprint.route('/form', methods=['GET'])
def get_forms():
    """
    Get all forms
    ---
    tags:
      - Dengue
    parameters:
      - name: user_id
        in: query
        type: string
        required: false
        description: The id of the user
    responses:
      200:
        description: Get all forms
    """
    rules = {}
    if api_input_check(['user_id'], request.args):
        rules['user_id'] = api_input_get(['user_id'], request.args)[0]
    dengues = Dengue.query.join(Building).filter_by(**rules).all()
    return Response.response('get forms successful', [dengue.as_dict() for dengue in dengues])


@dengue_blueprint.route('/form-download', methods=['GET'])
@authorization_required([0, 1, 2])
def download_forms():
    """
    download all forms
    ---
    tags:
      - Dengue
    parameters:
      - name: start_year_month
        in: query
        type: string
        required: true
        description: start date
      - name: end_year_month
        in: query
        type: string
        required: true
        description: end date
    responses:
      200:
        description: get restaurant_post stats success
      400:
        description: missing ['start_date', 'end_date'] query data
    """
    if not api_input_check(['start_date', 'end_date'], request.args):
        return Response.client_error("missing ['start_date', 'end_date'] query data")

    start_year_month, end_year_month = api_input_get(['start_date', 'end_date'], request.args)

    buildings = [building.chinese_name for building in Building.query.all()]
    time_range = [to_year_month(_) for _ in range(to_month(start_year_month) - 1, to_month(end_year_month))]

    df = pd.DataFrame("未填", index=buildings, columns=time_range)
    unfinished_df_map = {}
    for dengue, chinese_name in Dengue.query.join(Building).add_columns(Building.chinese_name).all():
        dengue_data_class = DengueDataClass(dengue)
        if chinese_name not in buildings or dengue.create_year_month not in time_range:
            continue
        if not dengue_data_class.finish:
            unfinished_df_map[(chinese_name, dengue.create_year_month)] = DengueDataClass(dengue).dengue_df
        df.loc[chinese_name, dengue.create_year_month] = dengue_data_class.finish

    wb = Workbook()
    ws = wb.active
    ws.title = '總覽'
    df.index.name = '建築物'
    df.reset_index(inplace=True)
    for r in dataframe_to_rows(df, index=False, header=True):
        ws.append(r)

    for dengue_chinese_name, unfinished_df in unfinished_df_map.items():
        ws = wb.create_sheet(title='_'.join(dengue_chinese_name))
        for r in dataframe_to_rows(unfinished_df, index=False, header=True):
            ws.append(r)

    excel_data = BytesIO()
    wb.save(excel_data)
    excel_data.seek(0)

    return send_file(
        excel_data,
        as_attachment=True,
        download_name=f'{start_year_month}_{end_year_month}_總覽.xlsx',
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )


@dengue_blueprint.route('/form-download/<int:form_id>', methods=['GET'])
@authorization_required([0, 1, 2])
def download_form(form_id):
    """
    download all forms
    ---
    tags:
      - Dengue
    parameters:
      - name: form_id
        in: path
        type: integer
        required: true
        description: form id
    responses:
      200:
        description: get restaurant_post stats success
      400:
        description: missing ['start_date', 'end_date'] query data
    """

    dengue = (Dengue.query
              .join(Building, Dengue.building_id == Building.id)
              .filter(Building.id == form_id)
              .add_columns(Building.chinese_name)
              .first())

    if dengue is None:
        return Response.not_found('form not found')

    dengue_data_class = DengueDataClass(dengue[0])
    chinese_name = dengue[1]

    df = dengue_data_class.dengue_df
    df.columns = ['問題', '回答', '子問題', '子問題回答']

    wb = Workbook()
    ws = wb.active
    column_map = {0: 'A', 1: 'B', 2: 'C', 3: 'D'}
    for col_idx, col_name in column_map.items():
        ws[col_name + '1'] = df.columns[col_idx]

    for row_idx in range(len(df)):
        for col_idx, col_name in column_map.items():
            ws[col_name + str(row_idx + 2)] = df.iloc[row_idx, col_idx]

    excel_data = BytesIO()
    wb.save(excel_data)
    excel_data.seek(0)

    return send_file(
        excel_data,
        as_attachment=True,
        download_name=f'{chinese_name}_{dengue_data_class.dengue.create_year_month}.xlsx',
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )


@dengue_blueprint.route('/form-status', methods=['GET'])
def get_form_status():
    """
    get form status
    ---
    tags:
      - Dengue
    responses:
      200:
        description: get form status success
    """
    current_month = datetime.now().month
    current_year = datetime.now().year

    buildings_name2id = {building.chinese_name: building.id for building in Building.query.all()}
    buildings = {building.id: "0" for building in Building.query.all()}
    time_range = [to_month(f'{current_year}-{current_month}')]

    for dengue, chinese_name in Dengue.query.join(Building).add_columns(Building.chinese_name).all():
        if to_month(dengue.create_year_month) not in time_range:
            continue

        buildings[buildings_name2id[chinese_name]] = "1"

    return Response.response("get form status successful", buildings)
